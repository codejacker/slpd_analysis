import pandas as pd
import numpy as np
import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def write_pivot_to_sheet(writer, sheet_name, pivot_df, start_row, title, filters, title_fill=None, start_col=1):
    """Writes a styled pivot table with a title and filters to a specific location on a sheet."""
    
    if sheet_name not in writer.book.sheetnames:
        ws = writer.book.create_sheet(sheet_name)
        writer.sheets[sheet_name] = ws
    else:
        ws = writer.sheets[sheet_name]

    if title:
        cell = ws.cell(row=start_row, column=start_col, value=title)
        cell.font = Font(bold=True, size=14)
        if title_fill:
            cell.fill = title_fill
        start_row += 2

    if filters:
        filter_start_row = start_row
        ws.cell(row=filter_start_row, column=start_col, value="Filters Used:").font = Font(bold=True)
        filter_row_offset = 1
        for key, value in filters.items():
            cell = ws.cell(row=filter_start_row + filter_row_offset, column=start_col, value=f"• {key}:")
            cell.font = Font(bold=True)
            
            if isinstance(value, list) and len(value) > 1:
                filter_row_offset += 1
                for item in value:
                    ws.cell(row=filter_start_row + filter_row_offset, column=start_col + 1, value=str(item))
                    filter_row_offset += 1
            else:
                ws.cell(row=filter_start_row + filter_row_offset, column=start_col + 1, value=str(value))
                filter_row_offset += 1
        start_row = filter_start_row + filter_row_offset

    pivot_start_row = start_row + 2
    pivot_df.to_excel(writer, sheet_name=sheet_name, startrow=pivot_start_row - 1, startcol=start_col - 1)
    
    lighter_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    header_rows = pivot_df.columns.nlevels
    data_rows = len(pivot_df)
    num_cols = len(pivot_df.columns) + pivot_df.index.nlevels

    for r in ws.iter_rows(min_row=pivot_start_row, max_row=pivot_start_row + header_rows + data_rows, min_col=start_col, max_col=start_col + num_cols - 1):
        for cell in r:
            cell.border = thin_border

    for r in range(pivot_start_row, pivot_start_row + header_rows):
        for c in range(start_col, start_col + num_cols):
            ws.cell(row=r, column=c).fill = lighter_blue_fill
            ws.cell(row=r, column=c).font = header_font

    if 'Grand Total' in pivot_df.index:
        try:
            grand_total_row_idx = pivot_df.index.get_loc('Grand Total')
            if isinstance(grand_total_row_idx, int):
                grand_total_row = pivot_start_row + header_rows + grand_total_row_idx
                for c in range(start_col, start_col + num_cols):
                    ws.cell(row=grand_total_row, column=c).fill = lighter_blue_fill
                    ws.cell(row=grand_total_row, column=c).font = header_font
        except (KeyError, TypeError):
            pass

    for i in range(num_cols):
        col_letter = get_column_letter(start_col + i)
        max_length = 0
        for r in range(pivot_start_row, pivot_start_row + header_rows + data_rows):
            cell = ws.cell(row=r, column=start_col + i)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    return ws.max_row

def create_lrc_cycle_table(writer, sheet_name, pvbe_df, ra_df, pivot_spec, start_row, all_cols, table_positions):
    import pandas as pd

    amount_col = all_cols['amount_col']
    date_col = all_cols['date_col']
    acc_change_col = all_cols['acc_change_col']
    proc_step_col = all_cols['proc_step_col']

    pvbe_df['Component'] = 'PVBE'
    ra_df['Component'] = 'RA'
    base_df = pd.concat([pvbe_df, ra_df])
    
    base_df[date_col] = pd.to_datetime(base_df[date_col], errors='coerce')
    base_df['Quarter'] = base_df[date_col].dt.to_period('Q')
    base_df['IsQuarterEnd'] = base_df[date_col].dt.is_quarter_end
    base_df['IsQuarterStart'] = base_df[date_col].dt.is_quarter_start
    
    quarter_end_dates = base_df[base_df['IsQuarterEnd']][date_col].drop_duplicates().sort_values()
    quarter_mapping = {date.to_period('Q'): date.strftime('%d/%m/%Y') for date in quarter_end_dates}
    quarters = sorted(quarter_mapping.keys())
    
    row_specs = [
        {'name': 'יתרת פתיחה', 'type': 'opening_balance'},
        {'name': 'עסק חדש', 'type': 'acc_filter', 'acc_code': 100, 'date_filter': 'quarter_start'},
        {'name': 'שחרור', 'type': 'acc_filter', 'acc_code': 405, 'date_filter': 'quarter_end'},
        {'name': 'תיאומים בהתאם לניסיון', 'type': 'acc_filter', 'acc_code': 505, 'date_filter': 'quarter_end', 'ra_zero': True},
        {'name': 'שינוי הנחות', 'type': 'acc_filter', 'acc_code': 600, 'date_filter': 'quarter_end'},
        {'name': 'שינוי ל LRR', 'type': 'empty'},
        {'name': 'זקיפה לCSM', 'type': 'csm_calculation'},
        {'name': 'הוצאות מימון', 'type': 'sum_finance'},
        {'name': 'צבירת ריבית', 'type': 'acc_filter', 'acc_code': 200, 'date_filter': 'quarter_end'},
        {'name': 'שינוי בריבית שוטפת', 'type': 'acc_filter', 'acc_code': 300, 'date_filter': 'quarter_end'},
        {'name': 'אינפלציה', 'type': 'acc_filter', 'acc_code': 601, 'date_filter': 'quarter_end'},
        {'name': 'יתרת סגירה ליום', 'type': 'sum_movements'},
        {'name': 'Check - תקין', 'type': 'validation_check'}
    ]
    
    result_data = []
    for i, quarter in enumerate(quarters):
        quarter_label = quarter_mapping.get(quarter, str(quarter))
        for row_spec in row_specs:
            row_data = {'Row': row_spec['name'], 'Quarter': quarter_label, 'PVBE': 0, 'RA': 0}
            if row_spec['type'] not in ['sum_finance', 'sum_movements', 'validation_check', 'empty', 'tbd']:
                df_slice = pd.DataFrame()
                if row_spec['type'] == 'opening_balance':
                    if i > 0:
                        df_slice = base_df[(base_df['Quarter'] == quarters[i-1]) & base_df['IsQuarterEnd']]
                    else:
                        df_slice = base_df[(base_df['Quarter'] == quarter) & base_df['IsQuarterStart']]
                elif row_spec['type'] == 'acc_filter':
                    df_slice = base_df[base_df[acc_change_col] == row_spec['acc_code']]
                    date_filter = row_spec['date_filter']
                    if date_filter == 'quarter_end':
                        df_slice = df_slice[df_slice['IsQuarterEnd'] & (df_slice['Quarter'] == quarter)]
                    elif date_filter == 'quarter_start':
                        df_slice = df_slice[df_slice['IsQuarterStart'] & (df_slice['Quarter'] == quarter)]
                elif row_spec['type'] == 'csm_calculation':
                    pvbe_csm, ra_csm = 0, 0
                    q_df = base_df[base_df['Quarter'] == quarter]
                    if i == 0:
                        c1 = (q_df['Component'] == 'PVBE') & q_df[proc_step_col].str.contains('Recognize Profit \(Prd Start - Bef. Chge\)', regex=True, na=False) & (q_df[acc_change_col] == 801) & q_df['IsQuarterStart']
                        c2 = (q_df['Component'] == 'PVBE') & q_df[proc_step_col].str.contains('Allocate \(Disclosure\) \(PE After Change\)', regex=True, na=False) & (q_df[acc_change_col] == 620) & q_df['IsQuarterEnd']
                        c3 = (q_df['Component'] == 'PVBE') & q_df[proc_step_col].str.contains('Recognize Profit \(PE/DE Before Change\)', regex=True, na=False) & (q_df[acc_change_col] == 410) & q_df['IsQuarterEnd']
                        pvbe_csm = q_df[c1 | c2 | c3][amount_col].sum()
                        c4 = (q_df['Component'] == 'RA') & q_df[proc_step_col].str.contains('Allocate \(Disclosure\)\(Per.St.- Aft.Chg.\)', regex=True, na=False) & (q_df[acc_change_col] == 120) & q_df['IsQuarterStart']
                        c5 = (q_df['Component'] == 'RA') & q_df[proc_step_col].str.contains('Allocate \(Disclosure\) \(PE After Change\)', regex=True, na=False) & (q_df[acc_change_col] == 620) & q_df['IsQuarterEnd']
                        ra_csm = q_df[c4 | c5][amount_col].sum()
                    else:
                        c1 = (q_df['Component'] == 'PVBE') & q_df[proc_step_col].str.contains('Allocate \(Disclosure\)\(Per.St.- Aft.Chg.\)', regex=True, na=False) & (q_df[acc_change_col] == 120) & q_df['IsQuarterStart']
                        pvbe_csm = q_df[c1][amount_col].sum()
                        c4 = (q_df['Component'] == 'RA') & q_df[proc_step_col].str.contains('Allocate \(Disclosure\)\(Per.St.- Aft.Chg.\)', regex=True, na=False) & (q_df[acc_change_col] == 120) & q_df['IsQuarterStart']
                        c5 = (q_df['Component'] == 'RA') & q_df[proc_step_col].str.contains('Allocate \(Disclosure\) \(PE After Change\)', regex=True, na=False) & (q_df[acc_change_col] == 620) & q_df['IsQuarterEnd']
                        ra_csm = q_df[c4 | c5][amount_col].sum()
                    row_data['PVBE'] = pvbe_csm
                    row_data['RA'] = ra_csm
                
                if not df_slice.empty:
                    row_data['PVBE'] = df_slice[df_slice['Component'] == 'PVBE'][amount_col].sum()
                    if not (row_spec.get('type') == 'acc_filter' and row_spec.get('ra_zero')):
                        row_data['RA'] = df_slice[df_slice['Component'] == 'RA'][amount_col].sum()

            result_data.append(row_data)

    result_df = pd.DataFrame(result_data)
    pivot_df = result_df.pivot_table(index='Row', columns='Quarter', values=['PVBE', 'RA'], aggfunc='sum', fill_value=0)
    
    if not pivot_df.empty:
        new_columns = [(val, q_label) for q in quarters for val in ['PVBE', 'RA'] if (val, (q_label := quarter_mapping.get(q, str(q)))) in pivot_df.columns]
        pivot_df = pivot_df[new_columns]
        finance_rows = ['צבירת ריבית', 'שינוי בריבית שוטפת', 'אינפלציה']
        pivot_df.loc['הוצאות מימון'] = pivot_df.loc[finance_rows].sum()
        closing_rows = ['יתרת פתיחה', 'עסק חדש', 'שחרור', 'תיאומים בהתאם לניסיון', 'שינוי הנחות', 'שינוי ל LRR', 'זקיפה לCSM', 'הוצאות מימון']
        pivot_df.loc['יתרת סגירה ליום'] = pivot_df.loc[closing_rows].sum()
        final_row_order = [spec['name'] for spec in row_specs]
        pivot_df = pivot_df.reindex(final_row_order)

    pivot_df.index.name = None
    table_positions[sheet_name].append((pivot_spec['title'], start_row))
    display_filters = {'Data Source': 'Based on Filtered & CRE 6000 tables'}
    return write_pivot_to_sheet(writer, sheet_name, pivot_df, start_row, pivot_spec['title'], display_filters) + 10

def create_lic_cycle_table(writer, sheet_name, filtered_out_df, filtered_in_df, pivot_spec, start_row, all_cols, table_positions):
    import pandas as pd

    amount_col = all_cols['amount_col']
    date_col = all_cols['date_col']
    acc_change_col = all_cols['acc_change_col']
    proc_step_col = all_cols['proc_step_col']

    filtered_out_df['Component'] = 'PVBE'
    filtered_in_df['Component'] = 'RA'
    base_df = pd.concat([filtered_out_df, filtered_in_df])
    
    base_df[date_col] = pd.to_datetime(base_df[date_col], errors='coerce')
    base_df['Quarter'] = base_df[date_col].dt.to_period('Q')
    base_df['IsQuarterEnd'] = base_df[date_col].dt.is_quarter_end
    base_df['IsQuarterStart'] = base_df[date_col].dt.is_quarter_start
    
    quarter_end_dates = base_df[base_df['IsQuarterEnd']][date_col].drop_duplicates().sort_values()
    quarter_mapping = {date.to_period('Q'): date.strftime('%d/%m/%Y') for date in quarter_end_dates}
    quarters = sorted(quarter_mapping.keys())
    
    row_specs = [
        {'name': 'יתרת פתיחה', 'type': 'opening_balance'},
        # Current Year Section
        {'name': 'תביעות והוצאות שירותי ביטוח אחרות שהתהוו', 'type': 'claims_incurred_current'},
        {'name': 'שחרור', 'type': 'acc_filter', 'acc_code': 405, 'date_filter': 'quarter_end'},
        {'name': 'תיאומים בהתאם לניסיון', 'type': 'acc_filter', 'acc_code': 505, 'date_filter': 'quarter_end'},
        {'name': 'שינוי הנחות', 'type': 'acc_filter', 'acc_code': 600, 'date_filter': 'quarter_end'},
        # Previous Year Section
        {'name': 'שינויים המתייחסים לשירותי עבר- תיאום להתחייבויות בגין תביעות שהתהוו', 'type': 'claims_incurred_past'},
        {'name': 'שחרור', 'type': 'acc_filter', 'acc_code': 405, 'date_filter': 'quarter_end', 'year_type': 'previous'},
        {'name': 'תיאומים בהתאם לניסיון', 'type': 'acc_filter', 'acc_code': 506, 'date_filter': 'quarter_end'},
        {'name': 'שינוי הנחות', 'type': 'acc_filter', 'acc_code': 600, 'date_filter': 'quarter_end', 'year_type': 'previous'},
        {'name': 'שינוי ל LRR', 'type': 'acc_filter', 'acc_code': 608, 'date_filter': 'quarter_end'},
        {'name': 'הוצאות מימון', 'type': 'sum_finance'},
        {'name': 'צבירת ריבית', 'type': 'acc_filter', 'acc_code': 200, 'date_filter': 'quarter_end'},
        {'name': 'שינוי בריבית שוטפת', 'type': 'acc_filter', 'acc_code': 300, 'date_filter': 'quarter_end'},
        {'name': 'אינפלציה', 'type': 'acc_filter', 'acc_code': 601, 'date_filter': 'quarter_end'},
        {'name': 'יתרת סגירה ליום', 'type': 'sum_movements'},
        {'name': 'Check - תקין', 'type': 'validation_check'}
    ]
    
    result_data = []
    for i, quarter in enumerate(quarters):
        quarter_label = quarter_mapping.get(quarter, str(quarter))
        for row_spec in row_specs:
            row_data = {'Row': row_spec['name'], 'Quarter': quarter_label, 'PVBE': 0, 'RA': 0}
            
            if row_spec['type'] not in ['sum_finance', 'sum_movements', 'validation_check', 'claims_incurred_current', 'claims_incurred_past']:
                df_slice = pd.DataFrame()
                if row_spec['type'] == 'opening_balance':
                    if i > 0:
                        df_slice = base_df[(base_df['Quarter'] == quarters[i-1]) & base_df['IsQuarterEnd']]
                    else:
                        df_slice = base_df[(base_df['Quarter'] == quarter) & base_df['IsQuarterStart']]
                elif row_spec['type'] == 'acc_filter':
                    df_slice = base_df[base_df[acc_change_col] == row_spec['acc_code']]
                    date_filter = row_spec['date_filter']
                    
                    # Handle current vs previous year logic
                    if row_spec.get('year_type') == 'previous':
                        if i > 0:
                            target_quarter = quarters[i-1]
                        else:
                            target_quarter = quarter  # fallback for first quarter
                    else:
                        target_quarter = quarter
                    
                    if date_filter == 'quarter_end':
                        df_slice = df_slice[df_slice['IsQuarterEnd'] & (df_slice['Quarter'] == target_quarter)]
                    elif date_filter == 'quarter_start':
                        df_slice = df_slice[df_slice['IsQuarterStart'] & (df_slice['Quarter'] == target_quarter)]
                
                if not df_slice.empty:
                    row_data['PVBE'] = df_slice[df_slice['Component'] == 'PVBE'][amount_col].sum()
                    row_data['RA'] = df_slice[df_slice['Component'] == 'RA'][amount_col].sum()
            
            elif row_spec['type'] == 'claims_incurred_current':
                # Sum of current year: שחרור + תיאומים בהתאם לניסיון + שינוי הנחות
                release_current = base_df[(base_df[acc_change_col] == 405) & base_df['IsQuarterEnd'] & (base_df['Quarter'] == quarter)]
                experience_current = base_df[(base_df[acc_change_col] == 505) & base_df['IsQuarterEnd'] & (base_df['Quarter'] == quarter)]
                discount_current = base_df[(base_df[acc_change_col] == 600) & base_df['IsQuarterEnd'] & (base_df['Quarter'] == quarter)]
                
                row_data['PVBE'] = (release_current[release_current['Component'] == 'PVBE'][amount_col].sum() + 
                                   experience_current[experience_current['Component'] == 'PVBE'][amount_col].sum() + 
                                   discount_current[discount_current['Component'] == 'PVBE'][amount_col].sum())
                row_data['RA'] = (release_current[release_current['Component'] == 'RA'][amount_col].sum() + 
                                 experience_current[experience_current['Component'] == 'RA'][amount_col].sum() + 
                                 discount_current[discount_current['Component'] == 'RA'][amount_col].sum())
            
            elif row_spec['type'] == 'claims_incurred_past':
                # Sum of previous year: שחרור + תיאומים בהתאם לניסיון + שינוי הנחות
                if i > 0:
                    prev_quarter = quarters[i-1]
                    release_past = base_df[(base_df[acc_change_col] == 405) & base_df['IsQuarterEnd'] & (base_df['Quarter'] == prev_quarter)]
                    experience_past = base_df[(base_df[acc_change_col] == 506) & base_df['IsQuarterEnd'] & (base_df['Quarter'] == prev_quarter)]
                    discount_past = base_df[(base_df[acc_change_col] == 600) & base_df['IsQuarterEnd'] & (base_df['Quarter'] == prev_quarter)]
                    
                    row_data['PVBE'] = (release_past[release_past['Component'] == 'PVBE'][amount_col].sum() + 
                                       experience_past[experience_past['Component'] == 'PVBE'][amount_col].sum() + 
                                       discount_past[discount_past['Component'] == 'PVBE'][amount_col].sum())
                    row_data['RA'] = (release_past[release_past['Component'] == 'RA'][amount_col].sum() + 
                                     experience_past[experience_past['Component'] == 'RA'][amount_col].sum() + 
                                     discount_past[discount_past['Component'] == 'RA'][amount_col].sum())

            result_data.append(row_data)

    result_df = pd.DataFrame(result_data)
    pivot_df = result_df.pivot_table(index='Row', columns='Quarter', values=['PVBE', 'RA'], aggfunc='sum', fill_value=0)
    
    if not pivot_df.empty:
        new_columns = [(val, q_label) for q in quarters for val in ['PVBE', 'RA'] if (val, (q_label := quarter_mapping.get(q, str(q)))) in pivot_df.columns]
        pivot_df = pivot_df[new_columns]
        
        # Calculate finance expenses sum
        finance_rows = ['צבירת ריבית', 'שינוי בריבית שוטפת', 'אינפלציה']
        pivot_df.loc['הוצאות מימון'] = pivot_df.loc[finance_rows].sum()
        
        # Calculate closing balance sum
        closing_rows = ['יתרת פתיחה', 'תביעות והוצאות שירותי ביטוח אחרות שהתהוו', 'שחרור', 'תיאומים בהתאם לניסיון', 'שינוי הנחות', 
                       'שינויים המתייחסים לשירותי עבר- תיאום להתחייבויות בגין תביעות שהתהוו', 'שינוי ל LRR', 'הוצאות מימון']
        pivot_df.loc['יתרת סגירה ליום'] = pivot_df.loc[closing_rows].sum()
        
        final_row_order = [spec['name'] for spec in row_specs]
        pivot_df = pivot_df.reindex(final_row_order)

    pivot_df.index.name = None
    table_positions[sheet_name].append((pivot_spec['title'], start_row))
    display_filters = {'Data Source': 'Based on Filtered Out & Filtered In tables'}
    return write_pivot_to_sheet(writer, sheet_name, pivot_df, start_row, pivot_spec['title'], display_filters) + 10

def create_csm_cycle_table(writer, sheet_name, csm_df, fv_df, pivot_spec, start_row, all_cols, table_positions):
    import pandas as pd

    amount_col = all_cols['amount_col']
    date_col = all_cols['date_col']
    acc_change_col = all_cols['acc_change_col']
    proc_step_col = all_cols['proc_step_col']

    # Use only CSM data for the cycle (as per user instruction to disregard F.V for columns)
    base_df = csm_df.copy()
    base_df['Component'] = 'CSM'
    
    base_df[date_col] = pd.to_datetime(base_df[date_col], errors='coerce')
    base_df['Quarter'] = base_df[date_col].dt.to_period('Q')
    base_df['IsQuarterEnd'] = base_df[date_col].dt.is_quarter_end
    base_df['IsQuarterStart'] = base_df[date_col].dt.is_quarter_start
    
    quarter_end_dates = base_df[base_df['IsQuarterEnd']][date_col].drop_duplicates().sort_values()
    quarter_mapping = {date.to_period('Q'): date.strftime('%d/%m/%Y') for date in quarter_end_dates}
    quarters = sorted(quarter_mapping.keys())
    
    row_specs = [
        {'name': 'יתרת פתיחה', 'type': 'opening_balance'},
        {'name': 'תיאומים בהתאם לניסיון', 'type': 'acc_filter', 'acc_code': 505, 'date_filter': 'quarter_end'},
        {'name': 'שינוי הנחות', 'type': 'acc_filter', 'acc_code': 600, 'date_filter': 'quarter_end'},
        {'name': 'הוצאות מימון', 'type': 'sum_finance'},
        {'name': 'צבירת ריבית', 'type': 'acc_filter', 'acc_code': 200, 'date_filter': 'quarter_end'},
        {'name': 'שינוי בריבית שוטפת', 'type': 'acc_filter', 'acc_code': 300, 'date_filter': 'quarter_end'},
        {'name': 'אינפלציה', 'type': 'acc_filter', 'acc_code': 601, 'date_filter': 'quarter_end'},
        {'name': 'CSM', 'type': 'acc_filter', 'acc_code': 410, 'date_filter': 'quarter_end'},
        {'name': 'יתרת סגירה ליום', 'type': 'sum_movements'},
        {'name': 'Check - תקין', 'type': 'validation_check'}
    ]
    
    result_data = []
    for i, quarter in enumerate(quarters):
        quarter_label = quarter_mapping.get(quarter, str(quarter))
        for row_spec in row_specs:
            row_data = {'Row': row_spec['name'], 'Quarter': quarter_label, 'CSM': 0}
            if row_spec['type'] not in ['sum_finance', 'sum_movements', 'validation_check']:
                df_slice = pd.DataFrame()
                if row_spec['type'] == 'opening_balance':
                    if i > 0:
                        df_slice = base_df[(base_df['Quarter'] == quarters[i-1]) & base_df['IsQuarterEnd']]
                    else:
                        df_slice = base_df[(base_df['Quarter'] == quarter) & base_df['IsQuarterStart']]
                elif row_spec['type'] == 'acc_filter':
                    df_slice = base_df[base_df[acc_change_col] == row_spec['acc_code']]
                    date_filter = row_spec['date_filter']
                    if date_filter == 'quarter_end':
                        df_slice = df_slice[df_slice['IsQuarterEnd'] & (df_slice['Quarter'] == quarter)]
                    elif date_filter == 'quarter_start':
                        df_slice = df_slice[df_slice['IsQuarterStart'] & (df_slice['Quarter'] == quarter)]
                
                if not df_slice.empty:
                    row_data['CSM'] = df_slice[amount_col].sum()

            result_data.append(row_data)

    result_df = pd.DataFrame(result_data)
    pivot_df = result_df.pivot_table(index='Row', columns='Quarter', values=['CSM'], aggfunc='sum', fill_value=0)
    
    if not pivot_df.empty:
        new_columns = [('CSM', q_label) for q in quarters if ('CSM', (q_label := quarter_mapping.get(q, str(q)))) in pivot_df.columns]
        pivot_df = pivot_df[new_columns]
        
        # Calculate finance expenses sum
        finance_rows = ['צבירת ריבית', 'שינוי בריבית שוטפת', 'אינפלציה']
        pivot_df.loc['הוצאות מימון'] = pivot_df.loc[finance_rows].sum()
        
        # Calculate closing balance sum
        closing_rows = ['יתרת פתיחה', 'תיאומים בהתאם לניסיון', 'שינוי הנחות', 'הוצאות מימון', 'CSM']
        pivot_df.loc['יתרת סגירה ליום'] = pivot_df.loc[closing_rows].sum()
        
        final_row_order = [spec['name'] for spec in row_specs]
        pivot_df = pivot_df.reindex(final_row_order)

    pivot_df.index.name = None
    table_positions[sheet_name].append((pivot_spec['title'], start_row))
    display_filters = {'Data Source': 'Based on מעגל CSM table'}
    return write_pivot_to_sheet(writer, sheet_name, pivot_df, start_row, pivot_spec['title'], display_filters) + 10

def get_filtered_df(df, spec, all_cols):
    filtered_df = df.copy()
    display_filters = {}
    filters = spec.get('filters', {})
    for col_name, values in filters.items():
        col_prop = all_cols[col_name]
        if col_name == 'sub_acc_col':
            filtered_df = filtered_df[filtered_df[col_prop].str.startswith(values)]
            display_filters[f"{col_name} (Starts With)"] = values
        elif col_name == 'proc_step_col' and spec.get('proc_step_filter') == 'not_contains':
            filtered_df = filtered_df[~filtered_df[col_prop].str.contains('|'.join([re.escape(v) for v in values]), case=False, na=False)]
            display_filters[f"{col_name} (Not Contains)"] = values
        elif col_name == 'cost_elem_col' and spec.get('cost_elem_filter') == 'not_contains':
             filtered_df = filtered_df[~filtered_df[col_prop].isin(values)]
             display_filters[f"{col_name} (Not In)"] = values
        elif col_name == 'cost_elem_col' and spec.get('cost_elem_filter') == 'in':
             filtered_df = filtered_df[filtered_df[col_prop].isin(values)]
             display_filters[col_name] = values
        elif col_name == 'coverage_id_col' and values == 'VFP_CONTAINS_FILTER':
            filtered_df = filtered_df[filtered_df[col_prop].astype(str).str.contains('VFP', case=False, na=False)]
            display_filters[f"{col_name} (Contains)"] = 'VFP'
        elif col_name == 'desc_gl_col' and spec.get('desc_gl_filter') == 'regex':
            pattern = values if isinstance(values, str) else '|'.join(values)
            filtered_df = filtered_df[filtered_df[col_prop].astype(str).str.contains(pattern, case=False, na=False, regex=True)]
            display_filters[f"{col_name} (Regex)"] = values
        else:
            filtered_df = filtered_df[filtered_df[col_prop].isin(values)]
            display_filters[col_name] = values
    return filtered_df, display_filters

def create_final_report(file_path, output_path):
    try:
        df = pd.read_excel(file_path, sheet_name='SLPD', header=0)
        all_cols = {
            'amount_col': 'Amount in Functional Currency', 'date_col': 'Posting Date',
            'class_col': 'Classification', 'cost_elem_col': 'Cost or Revenue Element',
            'gl_col': 'G/L Account', 'lifecycle_col': 'Subledger Account Lifecycle Stage',
            'sub_acc_col': 'Subledger Account', 'proc_step_col': 'Description Process Step ID',
            'loss_comp_col': 'Contributes to Loss Component', 'coverage_id_col': 'Coverage ID',
            'desc_gl_col': 'Description G/L Account', 'occ_year_col': 'Description Occurrence Year',
            'acc_change_col': 'Accounting Change'
        }
        for col in all_cols.values():
            if col not in df.columns:
                raise ValueError(f"Required column '{col}' not found.")
        df[all_cols['amount_col']] = pd.to_numeric(df[all_cols['amount_col']], errors='coerce').fillna(0)
        df[all_cols['sub_acc_col']] = df[all_cols['sub_acc_col']].astype(str)
        df[all_cols['date_col']] = df[all_cols['date_col']].astype(str)
        df[all_cols['acc_change_col']] = pd.to_numeric(df[all_cols['acc_change_col']], errors='coerce').fillna(0).astype(int)

        pivot_groups = {
            'LRC_VFA_Report': [
                {
                    'title': 'בדיקת סיווג רכיבי LRC לחשבונות GL הנכונים', 'filters': {'class_col': ['VFP'], 'loss_comp_col': [0], 'lifecycle_col': [0, 10], 'sub_acc_col': '1', 'proc_step_col': ['carry forward', 'release margin', 'value TC']}, 'proc_step_filter': 'not_contains', 'index': ['coverage_id_col'], 'columns': 'cost_elem_col', 'column_filter': ['6000', 'Z2002', 'Z4000', 'Z3100', 'Z1013', 'Z2004', 'Z1017', 'Z2005', 'Z2007', 'Z2012', 'Z2017', 'Z2008', 'Z5040', 'Z5050', 'Z6001']
                },
                {
                    'title': 'G/L Account Analysis', 'filters': {'class_col': ['VFP'], 'loss_comp_col': [0], 'proc_step_col': ['carry forward', 'release margin', 'value TC']}, 'proc_step_filter': 'not_contains', 'index': ['gl_col', 'desc_gl_col', 'cost_elem_col'], 'columns': 'date_col'
                },
                {
                    'title': 'בדיקת סבירות היוונים', 'filters': {'class_col': ['VFP'], 'loss_comp_col': [0], 'lifecycle_col': [0, 10], 'sub_acc_col': '1', 'proc_step_col': ['Capture (Central GAAP) (PE/DE Bef. Chg.)', 'Capture (Central GAAP) (PS - Bef. Chge)', 'Unwind & Release (PS - Before Change)', 'Unwind and Release (PE/DE Before Change)']}, 'proc_step_filter': 'isin', 'index': 'proc_step_col', 'columns': 'date_col'
                },
                {
                    'title': 'בדיקת סבירות RA', 'filters': {'class_col': ['VFP'], 'loss_comp_col': [0], 'lifecycle_col': [0, 10], 'sub_acc_col': '1', 'proc_step_col': ['carry forward', 'release margin', 'value TC'], 'cost_elem_col': ['6000', 'Z2002']}, 'proc_step_filter': 'not_contains', 'index': 'coverage_id_col', 'columns': 'date_col', 'title_color': '90EE90'
                },
                {
                    'layout': 'side_by_side',
                    'table1': {'id': 'pvbe_source_data', 'title': 'בדיקת סיווג רכיבי LRC - Filtered Cost Elements', 'filters': {'class_col': ['VFP'], 'loss_comp_col': [0], 'lifecycle_col': [0, 10], 'sub_acc_col': '1', 'cost_elem_col': ['6000', 'Z6001'], 'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)']}, 'cost_elem_filter': 'not_contains', 'proc_step_filter': 'not_contains', 'index': ['proc_step_col', 'acc_change_col'], 'columns': 'date_col', 'title_color': '90EE90'},
                    'table2': {'id': 'ra_source_data', 'title': 'בדיקת סיווג רכיבי LRC - CRE 6000 Only', 'filters': {'class_col': ['VFP'], 'loss_comp_col': [0], 'lifecycle_col': [0, 10], 'sub_acc_col': '1', 'cost_elem_col': ['6000'], 'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)']}, 'cost_elem_filter': 'in', 'proc_step_filter': 'not_contains', 'index': ['proc_step_col', 'acc_change_col'], 'columns': 'date_col', 'title_color': '90EE90'}
                },
                {
                    'title': 'מעגל LRC', 'type': 'custom_lrc_cycle'
                }
            ],
            'LIC_VFA': [
                {
                    'title': 'בדיקת סיווג רכיבי LIC לחשבונות GL הנכונים',
                    'filters': {
                        'lifecycle_col': [20, 50],
                        'sub_acc_col': '1',
                        'proc_step_col': ['carry forward'],
                        'coverage_id_col': 'VFP_CONTAINS_FILTER',
                        'cost_elem_col': ['ZR', 'CRES']
                    },
                    'proc_step_filter': 'not_contains',
                    'cost_elem_filter': 'not_contains',
                    'index': 'coverage_id_col',
                    'columns': 'cost_elem_col'
                },
                {
                    'title': 'G/L Account Analysis - Carry Forward VFP',
                    'filters': {
                        'class_col': ['VFP'],
                        'proc_step_col': ['carry forward'],
                        'desc_gl_col': ['LIC PVFCF RA - BS VFA', 'LIC PVFCF Claims - BS VFA', 'LIC ULAE- VFA', 'LIC Change Claims- Past Service - P&L VFA', 'LIC -InsFinExp Change in Inflation BE- P&L VFA', 'LIC -InsFinExp Change in Inflation BE - P&L PAA', 'LIC Change RA-Current Service - P&L VFA', 'LIC Change RA- Past Service - P&L VFA', 'LIC Change Claims- Current Service - P&L VFA', 'LIC Change ULAE Current Service - P&L VFA', 'LIC Change ULAE Past Service- P&L VFA']
                    },
                    'proc_step_filter': 'not_contains',
                    'index': ['gl_col', 'desc_gl_col'],
                    'columns': 'date_col'
                },
                {
                    'title': 'בדיקת סבירות היוונים',
                    'filters': {
                        'class_col': ['VFP'],
                        'lifecycle_col': [20, 50],
                        'sub_acc_col': '1',
                        'proc_step_col': ['Value TC (Ins. Contr.) (PE/DE Bef.Chg.)', 'Unwind and Release (PE/DE Before Change)', 'Capture (Central GAAP) (PE/DE Bef. Chg.)', 'Unwind & Release (PS - Before Change)', 'Capture (Central GAAP) (PS - Bef.Chge)']
                    },
                    'proc_step_filter': 'isin',
                    'index': ['proc_step_col', 'cost_elem_col'],
                    'columns': 'date_col'
                },
                {
                    'title': 'בדיקת סבירות RA',
                    'filters': {
                        'class_col': ['VFP'],
                        'lifecycle_col': [20, 50],
                        'sub_acc_col': '1',
                        'proc_step_col': ['Unwind and Release (PE/DE Before Change)', 'Capture (Central GAAP) (PE/DE Bef. Chg.)', 'Unwind & Release (PS - Before Change)', 'Capture (Central GAAP) (PS - Bef.Chge)']
                    },
                    'proc_step_filter': 'isin',
                    'index': 'coverage_id_col',
                    'columns': 'cost_elem_col',
                    'column_filter': ['6000', 'Z2002']
                },
                {
                    'layout': 'side_by_side',
                    'table1': {
                        'title': 'Filtered Out Cost Elements',
                        'filters': {
                            'class_col': ['VFP'],
                            'lifecycle_col': [20, 50],
                            'sub_acc_col': '1',
                            'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)'],
                            'cost_elem_col': ['6000', '3103', '7000', '7010', '7005', 'Z6001']
                        },
                        'proc_step_filter': 'not_contains',
                        'cost_elem_filter': 'not_contains',
                        'index': ['occ_year_col', 'proc_step_col', 'acc_change_col'],
                        'columns': 'date_col'
                    },
                    'table2': {
                        'title': 'Filtered In Cost Elements',
                        'filters': {
                            'class_col': ['VFP'],
                            'lifecycle_col': [20, 50],
                            'sub_acc_col': '1',
                            'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)'],
                            'cost_elem_col': ['6000']
                        },
                        'proc_step_filter': 'not_contains',
                        'cost_elem_filter': 'in',
                        'index': ['occ_year_col', 'proc_step_col', 'acc_change_col'],
                        'columns': 'date_col'
                    }
                },
                {
                    'title': 'מעגל LIC', 'type': 'custom_lic_cycle'
                }
            ],'LC_VFA': [
                {
                    'layout': 'side_by_side',
                    'table1': {
                        'title': 'בדיקת סיווג רכיבי LC לחשבונות GL הנכונים',
                        'filters': {
                            'class_col': ['VFP'],
                            'lifecycle_col': [10],
                            'sub_acc_col': '1',
                            'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)'],
                            'loss_comp_col': [1],
                            'cost_elem_col': ['6000', '3103', 'Z1012', 'Z4000', 'Z3100', 'Z1013', 'Z1017', 'ZR200', 'ZR100', 'ZR102', 'ZR202']
                        },
                        'proc_step_filter': 'not_contains',
                        'index': ['coverage_id_col'],
                        'columns': 'cost_elem_col',
                        'title_color': '90EE90'
                    },
                    'table2': {
                        'title': 'בדיקת סיווג רכיבי LC לחשבונות GL הנכונים',
                        'filters': {
                            'class_col': ['VFP'],
                            'lifecycle_col': [10],
                            'sub_acc_col': '1',
                            'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)'],
                            'loss_comp_col': [1],
                            'cost_elem_col': ['Z2002', 'Z2004', 'Z2005', 'Z2007', 'Z2012', 'Z2017']
                        },
                        'proc_step_filter': 'not_contains',
                        'index': ['coverage_id_col'],
                        'columns': 'cost_elem_col'
                    }
                },
                {
                    'title': 'G/L Account Analysis - LC',
                    'filters': {
                        'class_col': ['VFP'],
                        'loss_comp_col': [1],
                        'proc_step_col': ['Carry Forward', 'Release Margin', 'Value TC'],
                        'desc_gl_col': '^(LRC|LIC).*VFA'
                    },
                    'proc_step_filter': 'not_contains',
                    'desc_gl_filter': 'regex',
                    'index': ['gl_col', 'desc_gl_col'],
                    'columns': 'date_col'
                }
            ],
            'CSM_VFA': [
                {
                    'title': 'בדיקת סיווג רכיבי CSM לחשבונות GL הנכונים',
                    'filters': {
                        'class_col': ['VFP'],
                        'sub_acc_col': '1',
                        'cost_elem_col': ['7010'],
                        'proc_step_col': ['Carry Forward']
                    },
                    'proc_step_filter': 'not_contains',
                    'index': 'coverage_id_col',
                    'columns': None
                },
                {
                    'layout': 'side_by_side',
                    'table1': {
                        'title': 'מעגל CSM',
                        'filters': {
                            'cost_elem_col': ['7010'],
                            'sub_acc_col': '1',
                            'class_col': ['VFP'],
                            'loss_comp_col': [0],
                            'proc_step_col': ['Carry Forward', 'Release Margin(PE/DE Before Change)', 'Value TC(Ins. Contracts)(Period Start)']
                        },
                        'proc_step_filter': 'not_contains',
                        'index': ['proc_step_col', 'acc_change_col'],
                        'columns': 'date_col'
                    },
                    'table2': {
                        'title': 'מעגל F.V',
                        'filters': {
                            'cost_elem_col': ['Z6001'],
                            'sub_acc_col': '1',
                            'class_col': ['VFP'],
                            'proc_step_col': ['Carry Forward']
                        },
                        'proc_step_filter': 'not_contains',
                        'index': ['loss_comp_col', 'proc_step_col', 'acc_change_col', 'lifecycle_col'],
                        'columns': 'date_col'
                    }
                },
                {
                    'title': 'מעגל CSM', 'type': 'custom_csm_cycle'
                }
            ],
            'DAC': [
                {
                    'title': 'בדיקת מעגל DAC מתוך הריצות',
                    'filters': {
                        'class_col': ['VFP'],
                        'cost_elem_col': ['3103']
                    },
                    'index': ['cost_elem_col', 'acc_change_col'],
                    'columns': 'date_col'
                },
                {
                    'title': 'בדיקות מעגל DAC מתוך crez3100',
                    'filters': {
                        'class_col': ['VFP'],
                        'cost_elem_col': ['Z3100']
                    },
                    'index': ['cost_elem_col', 'acc_change_col'],
                    'columns': 'date_col'
                },
                {
                    'title': 'בדיקת מעגל DAC מתוך ה G/L',
                    'filters': {
                        'class_col': ['VFP'],
                        'desc_gl_col': ['Actual Acquisition Cost - P&L VFA', 'LRC Acq.Cost amortization Expenses P&L   VFA']
                    },
                    'index': ['gl_col', 'desc_gl_col'],
                    'columns': 'date_col'
                }
            ]

        }

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Source_Data', index=False)
            table_positions = {}

            for sheet_name, pivots in pivot_groups.items():
                current_row = 1
                table_positions[sheet_name] = []
                
                # Extract PVBE and RA data for LRC cycle if this is LRC_VFA_Report
                pvbe_df, ra_df = pd.DataFrame(), pd.DataFrame()
                if sheet_name == 'LRC_VFA_Report':
                    # Find the side_by_side specification for source data
                    for pivot_spec in pivots:
                        if pivot_spec.get('layout') == 'side_by_side':
                            pvbe_df, _ = get_filtered_df(df, pivot_spec['table1'], all_cols)
                            ra_df, _ = get_filtered_df(df, pivot_spec['table2'], all_cols)
                            break
                
                # Extract Filtered Out and Filtered In data for LIC cycle if this is LIC_VFA
                filtered_out_df, filtered_in_df = pd.DataFrame(), pd.DataFrame()
                if sheet_name == 'LIC_VFA':
                    # Find the side_by_side specification for source data
                    for pivot_spec in pivots:
                        if pivot_spec.get('layout') == 'side_by_side' and pivot_spec['table1']['title'] == 'Filtered Out Cost Elements':
                            filtered_out_df, _ = get_filtered_df(df, pivot_spec['table1'], all_cols)
                            filtered_in_df, _ = get_filtered_df(df, pivot_spec['table2'], all_cols)
                            break
                
                # Extract CSM and F.V data for CSM cycle if this is CSM_VFA
                csm_df, fv_df = pd.DataFrame(), pd.DataFrame()
                if sheet_name == 'CSM_VFA':
                    # Find the side_by_side specification for source data
                    for pivot_spec in pivots:
                        if pivot_spec.get('layout') == 'side_by_side' and pivot_spec['table1']['title'] == 'מעגל CSM':
                            csm_df, _ = get_filtered_df(df, pivot_spec['table1'], all_cols)
                            fv_df, _ = get_filtered_df(df, pivot_spec['table2'], all_cols)
                            break
                
                for pivot_spec in pivots:
                    if pivot_spec.get('type') == 'custom_lrc_cycle':
                        current_row = create_lrc_cycle_table(writer, sheet_name, pvbe_df, ra_df, pivot_spec, current_row, all_cols, table_positions)
                    elif pivot_spec.get('type') == 'custom_lic_cycle':
                        current_row = create_lic_cycle_table(writer, sheet_name, filtered_out_df, filtered_in_df, pivot_spec, current_row, all_cols, table_positions)
                    elif pivot_spec.get('type') == 'custom_csm_cycle':
                        current_row = create_csm_cycle_table(writer, sheet_name, csm_df, fv_df, pivot_spec, current_row, all_cols, table_positions)
                    elif pivot_spec.get('layout') == 'side_by_side':
                        spec1, spec2 = pivot_spec['table1'], pivot_spec['table2']
                        df1, d_filters1 = get_filtered_df(df, spec1, all_cols)
                        df2, d_filters2 = get_filtered_df(df, spec2, all_cols)
                        
                        pivot1_df = pd.pivot_table(df1, values=all_cols['amount_col'], index=[all_cols[i] for i in spec1['index']], columns=all_cols[spec1['columns']], aggfunc="sum", fill_value=0, margins=True, margins_name='Grand Total')
                        pivot2_df = pd.pivot_table(df2, values=all_cols['amount_col'], index=[all_cols[i] for i in spec2['index']], columns=all_cols[spec2['columns']], aggfunc="sum", fill_value=0, margins=True, margins_name='Grand Total')
                        
                        table_positions[sheet_name].append((spec1['title'], current_row))
                        row_after_1 = write_pivot_to_sheet(writer, sheet_name, pivot1_df, start_row=current_row, title=spec1['title'], filters=d_filters1)
                        start_col_2 = pivot1_df.shape[1] + 5
                        table_positions[sheet_name].append((spec2['title'], current_row))
                        row_after_2 = write_pivot_to_sheet(writer, sheet_name, pivot2_df, start_row=current_row, title=spec2['title'], filters=d_filters2, start_col=start_col_2)
                        current_row = max(row_after_1, row_after_2) + 10
                    else:
                        df_filtered, d_filters = get_filtered_df(df, pivot_spec, all_cols)
                        
                        index_cols = [all_cols[i] for i in pivot_spec['index']] if isinstance(pivot_spec['index'], list) else all_cols[pivot_spec['index']]
                        column_col = all_cols[pivot_spec['columns']] if pivot_spec.get('columns') else None
                        pivot_df = pd.pivot_table(df_filtered, values=all_cols['amount_col'], index=index_cols, columns=column_col, aggfunc="sum", fill_value=0, margins=True, margins_name='Grand Total')
                        
                        if 'column_filter' in pivot_spec:
                            existing_columns = [col for col in pivot_spec['column_filter'] if col in pivot_df.columns]
                            if existing_columns:
                                if 'Grand Total' in pivot_df.columns: existing_columns.append('Grand Total')
                                pivot_df = pivot_df[existing_columns]
                        
                        title_fill = PatternFill(start_color=pivot_spec['title_color'], end_color=pivot_spec['title_color'], fill_type="solid") if 'title_color' in pivot_spec else None
                        table_positions[sheet_name].append((pivot_spec['title'], current_row))
                        current_row = write_pivot_to_sheet(writer, sheet_name, pivot_df, start_row=current_row, title=pivot_spec['title'], filters=d_filters, title_fill=title_fill) + 10

            # Create table of contents
            toc_data = [
                ('VFP Checks', '', '', 'sheet_header'),
                ('בדיקת סיווג רכיבי LRC לחשבונות GL הנכונים - VFP', 'LRC_VFA_Report', 'המטרה לבדוק את כללי הגזירה של החשבונות המאזניים בlrc', 'table'),
                ('G/L Account Analysis - VFP', 'LRC_VFA_Report', 'המטרה לבדוק סבירות ההיוונים על capture עבור כל CRE בנפרד', 'table'),
                ('בדיקת סבירות היוונים - VFP', 'LRC_VFA_Report', 'בסיקת סבירות של רכיב הסיכון מתוך תביעות', 'table'),
                ('בדיקת סבירות RA - VFP', 'LRC_VFA_Report', 'הבדיקה עבור מעגל LRC', 'table'),
                ('מעגל LRC', 'LRC_VFA_Report', 'הבדיקה עבור מעגל LRC', 'table'),
                ('LC VFA Checks', '', '', 'sheet_header'),
                ('בדיקת סיווג רכיבי LC לחשבונות GL הנכונים', 'LC_VFA', 'המטרה לבדוק את כללי הגזירה של החשבונות המאזניים בlic', 'table'),
                ('CSM Checks', '', '', 'sheet_header'),
                ('בדיקת סיווג רכיבי CSM לחשבונות GL הנכונים', 'CSM_VFA', 'המטרה לבדוק את כללי הגזירה של החשבונות', 'table'),
                ('DAC Checks', '', '', 'sheet_header'),
                ('בדיקת מעגל DAC מתוך הריצות', 'DAC', 'המטרה לבדוק כי מעגל DAC מחושב מריצות נכונות', 'table'),
                ('בדיקות מעגל DAC מתוך crez3100', 'DAC', 'המטרה לבדוק כי כל הCRE שאמורים להיות נכללו במעגל', 'table'),
                ('בדיקת מעגל DAC מתוך ה G/L', 'DAC', 'המטרה לבדוק את כללי הגזירה של החשבונות במעגל', 'table'),
            ]

            # Create DataFrame
            toc_df = pd.DataFrame(toc_data, columns=['הבדיקה', 'לינק לבדיקה', 'הסבר', 'type'])
            toc_df_display = toc_df[['הבדיקה', 'לינק לבדיקה', 'הסבר']].copy()
            toc_df_display.to_excel(writer, sheet_name='ריכוז בדיקות', index=False)

            workbook = writer.book
            toc_sheet = workbook['ריכוז בדיקות']
            toc_sheet.sheet_view.rightToLeft = True

            # Apply formatting
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            link_font = Font(color="0000FF", underline="single")

            # Format headers
            for col in range(1, 4):
                cell = toc_sheet.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.font = header_font

            # Format data rows
            for idx, (check_name, sheet_link, explanation, row_type) in enumerate(toc_data, start=2):
                # Column A - Check name
                cell_a = toc_sheet.cell(row=idx, column=1, value=check_name)
                if row_type == 'sheet_header' or idx == len(toc_data) + 1:  # Last row gets same color as headers
                    cell_a.fill = green_fill
                    cell_a.font = header_font

                # Column B - Link
                cell_b = toc_sheet.cell(row=idx, column=2)
                if row_type == 'table' and sheet_link:
                    # Find the specific table position with exact title matching
                    table_row = 1  # Default to A1
                    if sheet_link in table_positions:
                        # Create a mapping of TOC names to actual table titles
                        title_mapping = {
                            'בדיקת סיווג רכיבי LRC לחשבונות GL הנכונים - VFP': 'בדיקת סיווג רכיבי LRC לחשבונות GL הנכונים',
                            'G/L Account Analysis - VFP': 'G/L Account Analysis',
                            'בדיקת סבירות היוונים - VFP': 'בדיקת סבירות היוונים',
                            'בדיקת סבירות RA - VFP': 'בדיקת סבירות RA',
                            'בדיקת סיווג רכיבי LRC - Filtered Cost Elements': 'בדיקת סיווג רכיבי LRC - Filtered Cost Elements',
                            'בדיקת סיווג רכיבי LRC - CRE 6000 Only': 'בדיקת סיווג רכיבי LRC - CRE 6000 Only',
                            'מעגל LRC': 'מעגל LRC',
                            'בדיקת סיווג רכיבי LIC לחשבונות GL הנכונים': 'בדיקת סיווג רכיבי LIC לחשבונות GL הנכונים',
                            'G/L Account Analysis - Carry Forward VFP': 'G/L Account Analysis - Carry Forward VFP',
                            'בדיקת היוונים עבור כל CRE': 'בדיקת סבירות היוונים',
                            'בדיקת סבירות RA': 'בדיקת סבירות RA',
                            'Filtered Out Cost Elements': 'Filtered Out Cost Elements',
                            'Filtered In Cost Elements': 'Filtered In Cost Elements',
                            'מעגל LIC': 'מעגל LIC',
                            'בדיקת סיווג רכיבי LC לחשבונות GL הנכונים': 'בדיקת סיווג רכיבי LC לחשבונות GL הנכונים',
                            'G/L Account Analysis - LC': 'G/L Account Analysis - LC',
                            'בדיקת סיווג רכיבי CSM לחשבונות GL הנכונים': 'בדיקת סיווג רכיבי CSM לחשבונות GL הנכונים',
                            'מעגל CSM': 'מעגל CSM',
                            'מעגל F.V': 'מעגל F.V',
                            'בדיקת מעגל DAC מתוך הריצות': 'בדיקת מעגל DAC מתוך הריצות',
                            'בדיקות מעגל DAC מתוך crez3100': 'בדיקות מעגל DAC מתוך crez3100',
                            'בדיקת מעגל DAC מתוך ה G/L': 'בדיקת מעגל DAC מתוך ה G/L'
                        }

                        target_title = title_mapping.get(check_name, check_name)
                        for title, row_pos in table_positions[sheet_link]:
                            if title == target_title:
                                table_row = row_pos
                                break

                    link = f"#\'{sheet_link}\'!A{table_row}\""
                    cell_b.value = check_name
                    cell_b.hyperlink = link
                    cell_b.font = link_font
                    if idx == len(toc_data) + 1:  # Last row gets same color formatting
                        cell_b.fill = green_fill
                        cell_b.font = Font(bold=True, color="000000", underline="single")
                
                # Column C - Explanation
                cell_c = toc_sheet.cell(row=idx, column=3, value=explanation)
                if idx == len(toc_data) + 1:  # Last row gets same color as headers
                    cell_c.fill = green_fill

            # Auto-fit columns
            for col in ['A', 'B', 'C']:
                toc_sheet.column_dimensions[col].width = 50

        print(f"\nSuccessfully created the report:\n{output_path}")

    except Exception as e:
        print(f"\nAn error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    input_file_path = filedialog.askopenfilename(title="Select the source Excel file", filetypes=(("Excel Files", "*.xlsx *.xls"), ("All files", "*.*")))
    if input_file_path:
        output_dir_path = filedialog.askdirectory(title="Select Output Folder")
        if output_dir_path:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            output_filename = f"final_report_{timestamp}.xlsx"
            full_output_path = os.path.join(output_dir_path, output_filename)
            print(f"\nInput file: {input_file_path}")
            print(f"Output will be saved as: {full_output_path}")
            create_final_report(input_file_path, full_output_path)